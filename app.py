"""
Bajet Sunuwai (बजेट सुनुवाई) — Agentic Civic Budget Formulation Prototype
=========================================================================

A single-file Streamlit prototype demonstrating a "Twin-Engine" agentic
architecture for Nepali local government budget formulation, bridging the
gap between Step 3 (Tole/Settlement selection) and Step 5 (Integrated
Program & Policy formulation) of the 7-step planning process mandated
under the Local Government Operation Act, 2074.

Engine A (Bottom-Up):  Ingests, translates and clusters unstructured,
                        multi-dialect citizen complaints into the 5
                        statutory thematic sectors.
Engine B (Top-Down):    "Thatha" data verification — cross-references
                        demographic matrices against Strategic Master
                        Plan Mandates and ring-fences baseline capital
                        works BEFORE citizen complaints are parsed.

Run with:  streamlit run app.py
"""

import io
from datetime import datetime

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SECTORS = [
    "Infrastructure",
    "Social",
    "Economic",
    "Environment/Agriculture",
    "Governance",
]

NAVY = "1B365D"
LIGHT_GREY = "F2F2F2"
ACCENT_GOLD = "C99A2E"

URGENCY_KEYWORDS = {
    5: ["आपतकालीन", "emergency", "urgent", "जोखिम", "collapse", "ढल्न", "danger"],
    4: ["तत्काल", "immediately", "बाढी", "flood", "epidemic", "महामारी"],
}

FISCAL_YEARS = ["2082/83", "2083/84", "2084/85"]

DEMO_COMPLAINTS = [
    {
        "Ward": 1,
        "Sector": "Infrastructure",
        "Complaint (raw, multi-dialect)": "गाउँको मुख्य सडक वर्षामा पूरै डुब्छ, तत्काल कल्भर्ट चाहियो।",
        "Urgency (1-5)": 4,
    },
    {
        "Ward": 2,
        "Sector": "Social",
        "Complaint (raw, multi-dialect)": "स्कुलमे शौचालय नहिये, बेटी सभ के बहुत दिक्कत होई छै।",
        "Urgency (1-5)": 3,
    },
    {
        "Ward": 3,
        "Sector": "Economic",
        "Complaint (raw, multi-dialect)": "हाट बजार क्षेत्रमा विद्युतीय पोल र प्रकाश व्यवस्था छैन।",
        "Urgency (1-5)": 2,
    },
    {
        "Ward": 4,
        "Sector": "Environment/Agriculture",
        "Complaint (raw, multi-dialect)": "सिंचाई नहर टुटल बा, धान के खेत सुखा रहल बा, आपतकालीन मर्मत चाही।",
        "Urgency (1-5)": 5,
    },
    {
        "Ward": 5,
        "Sector": "Governance",
        "Complaint (raw, multi-dialect)": "वडा कार्यालयमा नागरिकता सिफारिस पाउन एक हप्ता कुर्नुपर्छ, डिजिटलाइजेसन चाहियो।",
        "Urgency (1-5)": 2,
    },
    {
        "Ward": 6,
        "Sector": "Infrastructure",
        "Complaint (raw, multi-dialect)": "पुल भत्किसकेको छ, यात्रुहरु जोखिममा छन्, collapse हुने अवस्थामा छ।",
        "Urgency (1-5)": 5,
    },
]

# ---------------------------------------------------------------------------
# Agent System Prompts
# ---------------------------------------------------------------------------
# These are the exact system prompts that would be handed to an LLM
# orchestrator (e.g. LangGraph nodes) in production. In this offline
# prototype, ``simulate_ingestion_agent`` etc. approximate the same
# behaviour deterministically so the app runs without API keys.

INGESTION_TRANSLATION_AGENT_PROMPT = """
You are the INGESTION & TRANSLATION AGENT for Bajet Sunuwai, a civic budget
intelligence platform serving Nepali local governments.

INPUT: Raw, unstructured citizen grievances collected at Tole Bhela
(settlement-level ward assemblies). These inputs may be written in Nepali,
Maithili, Bhojpuri, Tharu, or code-mixed combinations of these, and are
frequently ungrammatical, abbreviated, or transliterated.

YOUR TASKS, IN ORDER:
1. NORMALIZE: Translate/transliterate the raw complaint into a clean,
   administrative-register English log entry suitable for a government
   file, while preserving all place names, infrastructure references and
   quantities exactly.
2. SECTOR TAG: Classify the complaint into exactly ONE of the five
   statutory thematic sectors mandated under the Local Government
   Operation Act, 2074:
      - Infrastructure
      - Social
      - Economic
      - Environment/Agriculture
      - Governance
3. URGENCY SCORE: Assign an urgency score from 1 (routine / long-term) to
   5 (life-safety / emergency), based on:
      - explicit danger/collapse/flood/epidemic language,
      - vulnerability of the affected population (children, elderly,
        disabled, pregnant women),
      - recurrence (has this been raised before without resolution?).
4. DEDUPLICATE: Flag near-duplicate entries from the same ward/sector so
   the Contextual Allocator Agent does not double-count urgency.
5. OUTPUT FORMAT: Return a structured JSON object per complaint:
   { "ward": int, "sector": str, "clean_text": str, "urgency": int,
     "duplicate_of": Optional[int], "language_detected": str }

CONSTRAINTS:
- Never invent facts not present in the raw complaint.
- Never resolve ambiguity by guessing a ward or sector — if truly
  ambiguous, tag sector as "Governance" (catch-all for administrative
  processing) and flag "needs_human_review": true.
- Preserve any Nepali fiscal-year or NPR figures verbatim.
"""

CONTEXTUAL_ALLOCATOR_AGENT_PROMPT = """
You are the CONTEXTUAL ALLOCATOR AGENT for Bajet Sunuwai. You receive:
  (a) the structured, sector-tagged citizen complaint log from the
      Ingestion & Translation Agent (Engine A — Bottom-Up), and
  (b) the municipality's demographic matrix and Strategic Master Plan
      Mandates (Engine B — Top-Down "Thatha" verification).

YOUR ALLOCATION ALGORITHM IS STRICTLY SEQUENTIAL — DO NOT REORDER:

STEP 1 — MANDATORY RING-FENCING (Top-Down, runs FIRST, before any
complaint is read):
  Cross-reference the demographic matrix against statutory service-level
  benchmarks. If a structural gap is detected (e.g. population exceeds
  40,000 within a health-service radius with zero secondary hospitals,
  or population exceeds the statutory threshold for a secondary/high
  school with no facility present), you MUST inject a baseline capital
  project for that gap and ring-fence its estimated cost from the Total
  Budget Ceiling BEFORE any citizen-driven allocation occurs. This
  guarantees that essential long-horizon infrastructure is never skipped
  merely because no individual citizen happened to file a complaint
  about it.

STEP 2 — REMAINING BUDGET CALCULATION:
  Remaining Budget = Total Budget Ceiling − Sum(Ring-Fenced Projects).
  If ring-fenced mandates exceed the ceiling, flag a CRITICAL fiscal
  infeasibility warning and cap ring-fencing at 60% of the ceiling.

STEP 3 — SECTOR-WEIGHTED DISTRIBUTION (Bottom-Up):
  Compute each sector's aggregate urgency weight = sum of urgency scores
  of all (deduplicated) complaints in that sector. Distribute the
  Remaining Budget across the 5 sectors proportionally to these weights.
  A sector with zero complaints still receives a statutory floor of 3% of
  the Remaining Budget to prevent total neglect.

STEP 4 — WITHIN-SECTOR PROJECT ALLOCATION:
  Within each sector's allotment, distribute funds across individual
  complaints proportionally to their urgency score, converting each
  complaint into a named line-item project with a ward reference.

OUTPUT: A single itemized project list with columns: Project Name,
Sector, Ward, Origin (Ring-Fenced Mandate | Citizen-Driven), Urgency,
Allocated Amount (NPR).
"""

INDEPENDENT_AUDITOR_AGENT_PROMPT = """
You are the INDEPENDENT AUDITOR AGENT for Bajet Sunuwai. You run AFTER the
Contextual Allocator Agent and BEFORE the budget is exported for tabling.
You have no authority to change allocations — only to flag them.

CHECK 1 — POLITICAL / WARD BIAS:
  Flag any single ward receiving more than 30% of the Total Budget
  Ceiling, unless fully explained by a ring-fenced Strategic Master Plan
  Mandate located in that ward.

CHECK 2 — SECTORAL STARVATION:
  Flag any statutory sector receiving less than its 3% statutory floor,
  or receiving 0 NPR despite unresolved high-urgency (4-5) complaints.

CHECK 3 — FUNDING MISMATCH:
  Flag any ring-fenced mandate whose injected cost exceeds 50% of the
  Total Budget Ceiling (fiscal infeasibility risk), and any citizen
  project whose allocated amount is disproportionate to its urgency
  score relative to peer projects in the same sector (>2x the sector's
  per-urgency-point average).

CHECK 4 — MANDATE INTEGRITY:
  Confirm every demographic trigger that fired in Engine B has a
  corresponding ring-fenced line item in the final export. A trigger
  that fired but produced no line item is a CRITICAL integrity failure.

OUTPUT: A verification log — one line per check, PASS or FLAGGED with a
plain-language reason — appended to the budget document as an audit
trail, never silently altered.
"""

# ---------------------------------------------------------------------------
# Engine A — Ingestion & Translation (simulated)
# ---------------------------------------------------------------------------


def simulate_ingestion_agent(df: pd.DataFrame) -> pd.DataFrame:
    """Approximates INGESTION_TRANSLATION_AGENT_PROMPT deterministically.

    Boosts the user-supplied urgency slider when emergency-signal
    keywords are present in the raw complaint text, flags likely
    duplicates within the same (ward, sector) pair, and produces a
    normalized administrative log line.
    """
    out = df.copy()
    out["Detected Urgency Boost"] = 0
    out["Needs Human Review"] = False
    out["Normalized Log Entry"] = ""

    seen_pairs = {}
    dup_flags = []

    for idx, row in out.iterrows():
        text = str(row.get("Complaint (raw, multi-dialect)", ""))
        base_urgency = int(row.get("Urgency (1-5)", 1))
        boost = 0
        for score, keywords in URGENCY_KEYWORDS.items():
            if any(k.lower() in text.lower() for k in keywords):
                boost = max(boost, score - base_urgency)
        out.at[idx, "Detected Urgency Boost"] = max(boost, 0)

        sector = row.get("Sector", "Governance")
        if sector not in SECTORS:
            out.at[idx, "Needs Human Review"] = True
            sector = "Governance"

        ward = row.get("Ward", 0)
        pair_key = (ward, sector)
        is_dup = pair_key in seen_pairs
        dup_flags.append(seen_pairs.get(pair_key, np.nan))
        seen_pairs[pair_key] = idx

        out.at[idx, "Normalized Log Entry"] = (
            f"[Ward {ward} | {sector}] {text.strip()} "
            f"(admin-log, urgency={min(base_urgency + boost, 5)})"
        )

    out["Duplicate Of (row idx)"] = dup_flags
    out["Final Urgency"] = (
        out["Urgency (1-5)"].astype(int) + out["Detected Urgency Boost"]
    ).clip(upper=5)
    return out


# ---------------------------------------------------------------------------
# Engine B — Contextual Allocator (Top-Down ring-fencing + Bottom-Up split)
# ---------------------------------------------------------------------------


def calculate_ringfenced_projects(
    population: int,
    hospital_access: str,
    school_access: str,
    budget_ceiling: float,
) -> list:
    """Engine B — 'Thatha' Data Verification.

    Injects mandatory Strategic Master Plan projects BEFORE any citizen
    complaint is parsed, whenever a statutory demographic trigger fires.
    """
    projects = []

    if population > 40_000 and hospital_access == "No Secondary Hospital":
        cost = min(round(budget_ceiling * 0.25, 2), 80_000_000)
        projects.append(
            {
                "Project Name": "Secondary Hospital Construction (Statutory Mandate)",
                "Sector": "Social",
                "Ward": "Municipality-Wide",
                "Origin": "Ring-Fenced Mandate",
                "Urgency": 5,
                "Allocated Amount (NPR)": cost,
                "Trigger": (
                    f"Population {population:,} > 40,000 AND zero secondary "
                    "hospitals within health-service radius"
                ),
            }
        )

    if population > 25_000 and school_access == "No Secondary/High School":
        cost = min(round(budget_ceiling * 0.12, 2), 40_000_000)
        projects.append(
            {
                "Project Name": "Secondary/High School Construction (Statutory Mandate)",
                "Sector": "Social",
                "Ward": "Municipality-Wide",
                "Origin": "Ring-Fenced Mandate",
                "Urgency": 5,
                "Allocated Amount (NPR)": cost,
                "Trigger": (
                    f"Population {population:,} > 25,000 AND no secondary/"
                    "high school facility present"
                ),
            }
        )

    return projects


def run_contextual_allocator(
    complaints_df: pd.DataFrame,
    population: int,
    hospital_access: str,
    school_access: str,
    budget_ceiling: float,
) -> tuple:
    """Full Contextual Allocator Agent pipeline (Steps 1-4)."""

    # STEP 1: Mandatory ring-fencing (Top-Down, runs first)
    ringfenced = calculate_ringfenced_projects(
        population, hospital_access, school_access, budget_ceiling
    )
    ringfenced_total = sum(p["Allocated Amount (NPR)"] for p in ringfenced)

    critical_warning = None
    if ringfenced_total > budget_ceiling * 0.6:
        capped_total = budget_ceiling * 0.6
        scale = capped_total / ringfenced_total if ringfenced_total else 0
        for p in ringfenced:
            p["Allocated Amount (NPR)"] = round(p["Allocated Amount (NPR)"] * scale, 2)
        ringfenced_total = capped_total
        critical_warning = (
            "CRITICAL: Ring-fenced statutory mandates exceeded the Total "
            "Budget Ceiling and were scaled down to a 60% cap for fiscal "
            "feasibility. Municipality should seek supplementary/"
            "intergovernmental fiscal transfer."
        )

    # STEP 2: Remaining budget
    remaining_budget = max(budget_ceiling - ringfenced_total, 0)

    # STEP 3: Sector-weighted distribution
    sector_weights = {s: 0.0 for s in SECTORS}
    for _, row in complaints_df.iterrows():
        sector = row["Sector"] if row["Sector"] in SECTORS else "Governance"
        sector_weights[sector] += float(row.get("Final Urgency", row.get("Urgency (1-5)", 1)))

    total_weight = sum(sector_weights.values())
    statutory_floor = 0.03 * remaining_budget
    sector_allocations = {}

    if total_weight <= 0:
        # No complaints at all: split evenly.
        for s in SECTORS:
            sector_allocations[s] = remaining_budget / len(SECTORS)
    else:
        # First pass: proportional split.
        raw_alloc = {
            s: (sector_weights[s] / total_weight) * remaining_budget for s in SECTORS
        }
        # Enforce statutory floor by clawing back proportionally from
        # sectors that are above the floor.
        below_floor = {s: v for s, v in raw_alloc.items() if v < statutory_floor}
        shortfall = sum(statutory_floor - v for v in below_floor.values())
        above_floor = {s: v for s, v in raw_alloc.items() if s not in below_floor}
        above_total = sum(above_floor.values())
        for s in SECTORS:
            if s in below_floor:
                sector_allocations[s] = statutory_floor
            else:
                claw = (raw_alloc[s] / above_total) * shortfall if above_total else 0
                sector_allocations[s] = max(raw_alloc[s] - claw, 0)

    # STEP 4: Within-sector, per-complaint allocation
    line_items = []
    for s in SECTORS:
        sector_rows = complaints_df[
            complaints_df["Sector"].apply(lambda x: x if x in SECTORS else "Governance") == s
        ]
        sector_budget = sector_allocations[s]
        sector_urgency_sum = sector_rows.get(
            "Final Urgency", sector_rows.get("Urgency (1-5)", pd.Series(dtype=float))
        ).sum()

        if len(sector_rows) == 0 or sector_urgency_sum <= 0:
            continue

        for _, row in sector_rows.iterrows():
            urgency = float(row.get("Final Urgency", row.get("Urgency (1-5)", 1)))
            share = urgency / sector_urgency_sum
            amount = round(sector_budget * share, 2)
            complaint_preview = str(row.get("Complaint (raw, multi-dialect)", ""))[:60]
            line_items.append(
                {
                    "Project Name": f"Ward {row['Ward']} — {s}: {complaint_preview}...",
                    "Sector": s,
                    "Ward": row["Ward"],
                    "Origin": "Citizen-Driven",
                    "Urgency": urgency,
                    "Allocated Amount (NPR)": amount,
                    "Trigger": "Citizen Tole Bhela complaint",
                }
            )

    all_projects = ringfenced + line_items
    return all_projects, sector_allocations, ringfenced_total, remaining_budget, critical_warning


# ---------------------------------------------------------------------------
# Independent Auditor Agent (simulated)
# ---------------------------------------------------------------------------


def run_independent_auditor(
    projects: list,
    budget_ceiling: float,
    population: int,
    hospital_access: str,
    school_access: str,
    critical_warning: str,
) -> list:
    log = []
    projects_df = pd.DataFrame(projects)

    # CHECK 1 — Political / Ward bias
    if not projects_df.empty:
        ward_totals = projects_df.groupby("Ward")["Allocated Amount (NPR)"].sum()
        flagged_wards = ward_totals[ward_totals > 0.30 * budget_ceiling]
        # Exclude municipality-wide ring-fenced mandates from ward-bias check
        flagged_wards = flagged_wards[flagged_wards.index != "Municipality-Wide"]
        if len(flagged_wards) > 0:
            for ward, amt in flagged_wards.items():
                log.append(
                    (
                        "FLAGGED",
                        f"Check 1 (Ward Bias): Ward {ward} receives NPR "
                        f"{amt:,.2f} ({amt / budget_ceiling:.1%} of ceiling) "
                        "— exceeds 30% threshold without a ring-fenced mandate "
                        "justification.",
                    )
                )
        else:
            log.append(("PASS", "Check 1 (Ward Bias): No ward exceeds the 30% concentration threshold."))
    else:
        log.append(("PASS", "Check 1 (Ward Bias): No project line items to evaluate."))

    # CHECK 2 — Sectoral starvation
    if not projects_df.empty:
        sector_totals = projects_df.groupby("Sector")["Allocated Amount (NPR)"].sum()
        starved = [s for s in SECTORS if sector_totals.get(s, 0) <= 0]
        if starved:
            log.append(
                (
                    "FLAGGED",
                    f"Check 2 (Sectoral Starvation): Sector(s) {', '.join(starved)} "
                    "received zero allocation.",
                )
            )
        else:
            log.append(("PASS", "Check 2 (Sectoral Starvation): All 5 statutory sectors received funding."))
    else:
        log.append(("FLAGGED", "Check 2 (Sectoral Starvation): No allocations generated at all."))

    # CHECK 3 — Funding mismatch
    if critical_warning:
        log.append(("FLAGGED", f"Check 3 (Funding Mismatch): {critical_warning}"))
    else:
        ringfenced_total = projects_df[projects_df["Origin"] == "Ring-Fenced Mandate"][
            "Allocated Amount (NPR)"
        ].sum() if not projects_df.empty else 0
        if ringfenced_total > 0.5 * budget_ceiling:
            log.append(
                (
                    "FLAGGED",
                    f"Check 3 (Funding Mismatch): Ring-fenced mandates total NPR "
                    f"{ringfenced_total:,.2f}, exceeding 50% of the Total Budget "
                    "Ceiling — fiscal infeasibility risk.",
                )
            )
        else:
            log.append(("PASS", "Check 3 (Funding Mismatch): Ring-fenced mandate cost is within safe fiscal bounds."))

    # CHECK 4 — Mandate integrity
    triggers_fired = []
    if population > 40_000 and hospital_access == "No Secondary Hospital":
        triggers_fired.append("Secondary Hospital Construction (Statutory Mandate)")
    if population > 25_000 and school_access == "No Secondary/High School":
        triggers_fired.append("Secondary/High School Construction (Statutory Mandate)")

    if triggers_fired:
        present_names = set(projects_df["Project Name"]) if not projects_df.empty else set()
        missing = [t for t in triggers_fired if t not in present_names]
        if missing:
            log.append(
                (
                    "FLAGGED",
                    "Check 4 (Mandate Integrity): CRITICAL — trigger(s) fired "
                    f"but no corresponding line item found: {', '.join(missing)}.",
                )
            )
        else:
            log.append(("PASS", "Check 4 (Mandate Integrity): All fired demographic triggers have matching ring-fenced line items."))
    else:
        log.append(("PASS", "Check 4 (Mandate Integrity): No demographic triggers fired this cycle."))

    return log


# ---------------------------------------------------------------------------
# Excel Export Engine
# ---------------------------------------------------------------------------


def generate_excel_report(
    municipality_name: str,
    fiscal_year: str,
    budget_ceiling: float,
    projects: list,
    sector_allocations: dict,
    audit_log: list,
    population: int,
    hospital_access: str,
    school_access: str,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Integrated Budget"

    navy_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    grey_fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
    gold_fill = PatternFill(start_color=ACCENT_GOLD, end_color=ACCENT_GOLD, fill_type="solid")
    white_bold = Font(color="FFFFFF", bold=True, size=12)
    white_bold_small = Font(color="FFFFFF", bold=True, size=10)
    title_font = Font(color=NAVY, bold=True, size=16)
    subtitle_font = Font(color="404040", italic=True, size=10)
    thin_side = Side(style="thin", color="B0B0B0")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    currency_fmt = '#,##0.00 "NPR"'

    headers = ["Project Name", "Sector", "Ward", "Origin", "Urgency", "Allocated Amount (NPR)"]
    n_cols = len(headers)

    # ---- Title block ----
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1, value=f"{municipality_name} — Integrated Program & Budget (FY {fiscal_year})")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    generated = ws.cell(
        row=2,
        column=1,
        value=(
            f"Generated by Bajet Sunuwai Agentic Engine | "
            f"{datetime.now().strftime('%Y-%m-%d %H:%M')} | "
            f"Total Budget Ceiling: NPR {budget_ceiling:,.2f}"
        ),
    )
    generated.font = subtitle_font
    generated.alignment = Alignment(horizontal="center")

    context_lines = [
        f"Population Input: {population:,}",
        f"Hospital Access Status: {hospital_access}",
        f"School Access Status: {school_access}",
    ]
    if population > 40_000 and hospital_access == "No Secondary Hospital":
        context_lines.append("Demographic Trigger: Mandatory Hospital Construction Enforced")
    if population > 25_000 and school_access == "No Secondary/High School":
        context_lines.append("Demographic Trigger: Mandatory School Construction Enforced")

    row_cursor = 3
    for line in context_lines:
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=n_cols)
        c = ws.cell(row=row_cursor, column=1, value=f"• {line}")
        c.font = Font(color=NAVY, bold=True, size=10)
        c.alignment = Alignment(horizontal="left")
        c.fill = grey_fill
        row_cursor += 1

    row_cursor += 1  # spacer

    # ---- Sector summary block ----
    ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=n_cols)
    sec_title = ws.cell(row=row_cursor, column=1, value="Sector-Wise Ceiling Distribution")
    sec_title.font = white_bold_small
    sec_title.fill = gold_fill
    sec_title.alignment = Alignment(horizontal="left", vertical="center")
    row_cursor += 1

    for sector, amount in sector_allocations.items():
        ws.cell(row=row_cursor, column=1, value=sector).border = thin_border
        amt_cell = ws.cell(row=row_cursor, column=2, value=round(amount, 2))
        amt_cell.number_format = currency_fmt
        amt_cell.border = thin_border
        for col in range(3, n_cols + 1):
            ws.cell(row=row_cursor, column=col).border = thin_border
        row_cursor += 1

    row_cursor += 1  # spacer

    # ---- Project table header ----
    header_row = row_cursor
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.fill = navy_fill
        cell.font = white_bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[header_row].height = 24

    # ---- Project rows ----
    data_start = header_row + 1
    r = data_start
    for p in projects:
        ws.cell(row=r, column=1, value=p.get("Project Name", "")).border = thin_border
        ws.cell(row=r, column=2, value=p.get("Sector", "")).border = thin_border
        ws.cell(row=r, column=3, value=str(p.get("Ward", ""))).border = thin_border
        origin_cell = ws.cell(row=r, column=4, value=p.get("Origin", ""))
        origin_cell.border = thin_border
        if p.get("Origin") == "Ring-Fenced Mandate":
            origin_cell.font = Font(color="B7791F", bold=True)
        ws.cell(row=r, column=5, value=p.get("Urgency", "")).border = thin_border
        amt_cell = ws.cell(row=r, column=6, value=p.get("Allocated Amount (NPR)", 0))
        amt_cell.number_format = currency_fmt
        amt_cell.border = thin_border
        if r % 2 == 0:
            for col in range(1, n_cols + 1):
                ws.cell(row=r, column=col).fill = grey_fill
        r += 1
    data_end = r - 1

    # ---- SUM formula row ----
    sum_row = r
    ws.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=5)
    total_label = ws.cell(row=sum_row, column=1, value="TOTAL ALLOCATED BUDGET")
    total_label.font = white_bold
    total_label.fill = navy_fill
    total_label.alignment = Alignment(horizontal="right", vertical="center")
    for col in range(1, 6):
        ws.cell(row=sum_row, column=col).fill = navy_fill
        ws.cell(row=sum_row, column=col).border = thin_border

    total_col_letter = get_column_letter(6)
    total_cell = ws.cell(
        row=sum_row,
        column=6,
        value=f"=SUM({total_col_letter}{data_start}:{total_col_letter}{data_end})" if data_end >= data_start else 0,
    )
    total_cell.number_format = currency_fmt
    total_cell.font = white_bold
    total_cell.fill = navy_fill
    total_cell.border = thin_border
    ws.row_dimensions[sum_row].height = 22

    remainder_row = sum_row + 1
    ws.merge_cells(start_row=remainder_row, start_column=1, end_row=remainder_row, end_column=5)
    rem_label = ws.cell(row=remainder_row, column=1, value="TOTAL BUDGET CEILING (NPR)")
    rem_label.font = Font(bold=True, color=NAVY)
    rem_label.alignment = Alignment(horizontal="right")
    ceiling_cell = ws.cell(row=remainder_row, column=6, value=budget_ceiling)
    ceiling_cell.number_format = currency_fmt
    ceiling_cell.font = Font(bold=True, color=NAVY)

    # ---- Audit trail sheet ----
    audit_ws = wb.create_sheet("Independent Audit Log")
    audit_ws.merge_cells("A1:C1")
    audit_title = audit_ws.cell(row=1, column=1, value=f"Independent Auditor Agent — Verification Log ({municipality_name}, FY {fiscal_year})")
    audit_title.font = title_font
    audit_headers = ["Status", "Finding", "Timestamp"]
    for col_idx, h in enumerate(audit_headers, start=1):
        c = audit_ws.cell(row=3, column=col_idx, value=h)
        c.fill = navy_fill
        c.font = white_bold
        c.border = thin_border
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    for i, (status, msg) in enumerate(audit_log, start=4):
        status_cell = audit_ws.cell(row=i, column=1, value=status)
        status_cell.font = Font(
            bold=True, color=("C0392B" if status == "FLAGGED" else "1E8449")
        )
        status_cell.border = thin_border
        msg_cell = audit_ws.cell(row=i, column=2, value=msg)
        msg_cell.alignment = Alignment(wrap_text=True, vertical="top")
        msg_cell.border = thin_border
        audit_ws.cell(row=i, column=3, value=ts).border = thin_border
    audit_ws.column_dimensions["A"].width = 12
    audit_ws.column_dimensions["B"].width = 90
    audit_ws.column_dimensions["C"].width = 18

    # ---- Column widths (main sheet) ----
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 24
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------


def main():
    st.set_page_config(
        page_title="Bajet Sunuwai — Agentic Budget Formulation",
        page_icon="🏛️",
        layout="wide",
    )

    st.title("🏛️ बजेट सुनुवाई — Bajet Sunuwai")
    st.caption(
        "Agentic Twin-Engine platform bridging Tole Bhela grassroots inputs "
        "(Engine A) with statutory demographic mandates (Engine B) into an "
        "Integrated Program & Budget, per the Local Government Operation "
        "Act, 2074."
    )

    # ---- Sidebar: Global Parameters ----
    with st.sidebar:
        st.header("⚙️ Global Parameters")
        municipality_name = st.text_input("Municipality Name", value="Nagarain Municipality")
        fiscal_year = st.selectbox("Fiscal Year", FISCAL_YEARS, index=0)
        budget_ceiling = st.number_input(
            "Total Budget Ceiling (NPR)",
            min_value=1_000_000.0,
            max_value=10_000_000_000.0,
            value=250_000_000.0,
            step=1_000_000.0,
            format="%.2f",
        )

        st.divider()
        st.header("🧭 Demographic Parameters (Engine B — 'Thatha')")
        st.caption("Adjust these to test the Top-Down mandate-injection trigger.")
        population = st.slider("Municipal Population", min_value=5_000, max_value=150_000, value=45_000, step=1_000)
        hospital_access = st.selectbox(
            "Hospital Access Status",
            ["Has Secondary Hospital", "No Secondary Hospital"],
            index=1,
        )
        school_access = st.selectbox(
            "Secondary/High School Access Status",
            ["Has Secondary/High School", "No Secondary/High School"],
            index=0,
        )

        st.divider()
        with st.expander("📜 View Agent System Prompts"):
            st.markdown("**🤖 Ingestion & Translation Agent**")
            st.code(INGESTION_TRANSLATION_AGENT_PROMPT, language="text")
            st.markdown("**🤖 Contextual Allocator Agent**")
            st.code(CONTEXTUAL_ALLOCATOR_AGENT_PROMPT, language="text")
            st.markdown("**🤖 Independent Auditor Agent**")
            st.code(INDEPENDENT_AUDITOR_AGENT_PROMPT, language="text")

    # ---- Trigger status banner ----
    trigger_msgs = []
    if population > 40_000 and hospital_access == "No Secondary Hospital":
        trigger_msgs.append("🏥 Hospital construction mandate ARMED (population > 40,000, no secondary hospital).")
    if population > 25_000 and school_access == "No Secondary/High School":
        trigger_msgs.append("🏫 School construction mandate ARMED (population > 25,000, no secondary/high school).")
    if trigger_msgs:
        st.warning("  \n".join(trigger_msgs))
    else:
        st.success("No demographic mandate triggers currently armed. Adjust parameters in the sidebar to test Engine B.")

    st.subheader("📝 Step 3 → Step 4: Unstructured Citizen Requests (Tole Bhela Inputs)")
    st.caption(
        "Simulates raw, multi-dialect complaints collected at settlement-level ward "
        "assemblies. Edit, add, or delete rows to test Engine A's ingestion pipeline."
    )

    if "complaints_df" not in st.session_state:
        st.session_state.complaints_df = pd.DataFrame(DEMO_COMPLAINTS)

    edited_df = st.data_editor(
        st.session_state.complaints_df,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Ward": st.column_config.NumberColumn("Ward", min_value=1, max_value=20, step=1),
            "Sector": st.column_config.SelectboxColumn("Sector", options=SECTORS),
            "Complaint (raw, multi-dialect)": st.column_config.TextColumn(
                "Complaint (raw, multi-dialect)", width="large"
            ),
            "Urgency (1-5)": st.column_config.SliderColumn(
                "Baseline Urgency", min_value=1, max_value=5, step=1
            ),
        },
        key="complaints_editor",
    )
    st.session_state.complaints_df = edited_df

    run = st.button("🚀 Run Agentic Pipeline", type="primary", use_container_width=True)

    if run:
        if edited_df.empty:
            st.error("Add at least one citizen complaint row before running the pipeline.")
            return

        with st.spinner("Engine A: Ingestion & Translation Agent processing multi-dialect inputs..."):
            ingested_df = simulate_ingestion_agent(edited_df)

        st.subheader("🤖 Engine A Output — Ingestion & Translation Agent")
        st.dataframe(
            ingested_df[
                [
                    "Ward",
                    "Sector",
                    "Normalized Log Entry",
                    "Final Urgency",
                    "Needs Human Review",
                    "Duplicate Of (row idx)",
                ]
            ],
            use_container_width=True,
        )

        with st.spinner("Engine B + Contextual Allocator: ring-fencing mandates and distributing budget..."):
            (
                projects,
                sector_allocations,
                ringfenced_total,
                remaining_budget,
                critical_warning,
            ) = run_contextual_allocator(
                ingested_df, population, hospital_access, school_access, budget_ceiling
            )

        if critical_warning:
            st.error(critical_warning)

        col1, col2, col3 = st.columns(3)
        col1.metric("Total Budget Ceiling", f"NPR {budget_ceiling:,.0f}")
        col2.metric("Ring-Fenced (Mandates)", f"NPR {ringfenced_total:,.0f}")
        col3.metric("Remaining (Citizen-Driven)", f"NPR {remaining_budget:,.0f}")

        st.subheader("🤖 Engine B — Contextual Allocator Agent: Integrated Project List")
        projects_df = pd.DataFrame(projects)
        st.dataframe(projects_df, use_container_width=True)

        st.subheader("📊 Sector-Wise Ceiling Distribution")
        sector_df = pd.DataFrame(
            {"Sector": list(sector_allocations.keys()), "Allocated (NPR)": list(sector_allocations.values())}
        )
        st.bar_chart(sector_df.set_index("Sector"))

        with st.spinner("Independent Auditor Agent verifying allocations..."):
            audit_log = run_independent_auditor(
                projects, budget_ceiling, population, hospital_access, school_access, critical_warning
            )

        st.subheader("🤖 Independent Auditor Agent — Verification Log")
        for status, msg in audit_log:
            if status == "PASS":
                st.success(f"✅ {msg}")
            else:
                st.error(f"🚩 {msg}")

        excel_bytes = generate_excel_report(
            municipality_name,
            fiscal_year,
            budget_ceiling,
            projects,
            sector_allocations,
            audit_log,
            population,
            hospital_access,
            school_access,
        )

        st.subheader("📥 Export Integrated Program & Budget")
        st.download_button(
            label="Download Presentation-Ready Excel Report (.xlsx)",
            data=excel_bytes,
            file_name=f"{municipality_name.replace(' ', '_')}_Budget_FY{fiscal_year.replace('/', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        st.session_state["last_projects"] = projects
        st.session_state["last_audit_log"] = audit_log


if __name__ == "__main__":
    main()
